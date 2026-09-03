"""Input.py."""

from pathlib import Path

import numpy as np
import openpyxl

from reactormodels.models import (
    FreundlichIsotherm,
    LangmuirIsotherm,
    LinearIsotherm,
)
from reactormodels.properties import (
    Breakthrough,
    Chemical,
    Column,
    Media,
    Water,
)


def normalize_name(name: str) -> str:
    """Normalize chemical names for matching."""
    return "".join(str(name).split()).casefold()


def _load_properties(sheet):
    """Create Water, Media, and Column objects."""
    water_parameters = {}
    media_parameters = {}
    column_parameters = {}

    sections = {
        "water": water_parameters,
        "media": media_parameters,
        "column": column_parameters,
    }

    current_section = None

    for parameter, value, *_ in sheet.iter_rows(values_only=True):
        if parameter is None:
            continue

        parameter_name = str(parameter).strip().rstrip(":")
        key = parameter_name.casefold()

        if key in sections:
            current_section = sections[key]
            continue

        if key in {
            "parameter",
            "parameters",
            "experiment_id",
        }:
            continue

        if current_section is not None and value is not None:
            current_section[parameter_name] = value

    water = Water(**water_parameters)
    media = Media(**media_parameters)

    column = Column(
        water=water,
        media=media,
        **column_parameters,
    )

    return water, media, column


def _load_chemicals(sheet, compound_names):
    """Create Chemical objects used in the breakthrough data."""
    chemicals = {}

    normalized_compounds = {normalize_name(name) for name in compound_names}

    parameter_names = {
        column_number: str(sheet.cell(row=1, column=column_number).value).strip()
        for column_number in range(2, sheet.max_column + 1)
        if sheet.cell(row=1, column=column_number).value is not None
    }

    for row in range(2, sheet.max_row + 1):
        chemical_name = sheet.cell(
            row=row,
            column=1,
        ).value

        if chemical_name is None:
            continue

        normalized_name = normalize_name(chemical_name)

        if normalized_name not in normalized_compounds:
            continue

        chemical_parameters = {}

        for column_number, parameter_name in parameter_names.items():
            value = sheet.cell(
                row=row,
                column=column_number,
            ).value

            if value is not None:
                chemical_parameters[parameter_name] = value

        chemicals[normalized_name] = Chemical(
            name=str(chemical_name).strip(),
            **chemical_parameters,
        )

    return chemicals


def _load_breakthroughs(
    workbook,
    sheet_name,
    chemicals,
    column,
):
    """Create Breakthrough objects."""
    parameter_sheet = workbook["breakthrough_parameters"]
    feed_sheet = workbook["feed_concentration"]
    selected_sheet = workbook[sheet_name]
    initial_sheet = workbook["initial_concentration"]

    breakthrough_parameters = {}

    for parameter, value, *_ in parameter_sheet.iter_rows(values_only=True):
        if parameter is None or value is None:
            continue

        parameter_name = str(parameter).strip()

        if parameter_name.casefold() in {
            "breakthrough",
            "parameter",
            "parameters",
        }:
            continue

        breakthrough_parameters[parameter_name] = value

    feed_columns = {
        normalize_name(feed_sheet.cell(1, column_number).value): column_number
        for column_number in range(3, feed_sheet.max_column + 1)
        if feed_sheet.cell(1, column_number).value is not None
    }

    initial_columns = {
        normalize_name(initial_sheet.cell(1, column_number).value): column_number
        for column_number in range(
            3,
            initial_sheet.max_column + 1,
        )
        if initial_sheet.cell(1, column_number).value is not None
    }

    breakthroughs = {}

    for column_number in range(
        3,
        selected_sheet.max_column + 1,
    ):
        chemical_name = selected_sheet.cell(
            row=1,
            column=column_number,
        ).value

        if chemical_name is None:
            continue

        normalized_name = normalize_name(chemical_name)

        if normalized_name not in chemicals:
            raise ValueError(f"{chemical_name!r} is not in the chemical sheet.")

        selected_concentrations = []
        bed_volumes = []
        time = []

        for row in range(2, selected_sheet.max_row + 1):
            concentration = selected_sheet.cell(
                row=row,
                column=column_number,
            ).value

            if concentration is None:
                continue

            selected_concentrations.append(concentration)

            bed_volumes.append(
                selected_sheet.cell(
                    row=row,
                    column=1,
                ).value
            )

            time.append(
                selected_sheet.cell(
                    row=row,
                    column=2,
                ).value
            )

        feed_column = feed_columns[normalized_name]

        feed_concentrations = [
            feed_sheet.cell(
                row=row,
                column=feed_column,
            ).value
            for row in range(2, feed_sheet.max_row + 1)
            if feed_sheet.cell(
                row=row,
                column=feed_column,
            ).value
            is not None
        ]

        breakthrough_kwargs = {
            "chemical": chemicals[normalized_name],
            "column": column,
            "feed_concentrations": np.asarray(
                feed_concentrations,
                dtype=float,
            ),
            "bed_volumes": np.asarray(
                bed_volumes,
                dtype=float,
            ),
            "time": np.asarray(
                time,
                dtype=float,
            ),
            **breakthrough_parameters,
        }

        if sheet_name == "effluent_concentration":
            breakthrough_kwargs["effluent_concentrations"] = np.asarray(
                selected_concentrations,
                dtype=float,
            )

        if normalized_name in initial_columns:
            initial_concentration = initial_sheet.cell(
                row=2,
                column=initial_columns[normalized_name],
            ).value

            if initial_concentration is not None:
                breakthrough_kwargs["initial_concentration"] = initial_concentration

        breakthroughs[normalized_name] = Breakthrough(**breakthrough_kwargs)

    return breakthroughs


def _load_isotherms(
    sheet,
    compound_names,
):
    """Create isotherm objects from supplied parameters."""
    isotherms = {}

    normalized_compounds = {normalize_name(name) for name in compound_names}

    for row in range(4, sheet.max_row + 1):
        chemical_name = sheet.cell(
            row=row,
            column=1,
        ).value

        if chemical_name is None:
            continue

        normalized_name = normalize_name(chemical_name)

        if normalized_name not in normalized_compounds:
            continue

        chemical_isotherms = {}

        # Linear: column B
        linear_k = sheet.cell(
            row=row,
            column=2,
        ).value

        if linear_k is not None:
            chemical_isotherms["linear"] = LinearIsotherm(K=linear_k)

        # Freundlich: columns C-D
        freundlich_k = sheet.cell(
            row=row,
            column=3,
        ).value

        one_over_n = sheet.cell(
            row=row,
            column=4,
        ).value

        if freundlich_k is not None and one_over_n is not None:
            chemical_isotherms["freundlich"] = FreundlichIsotherm(
                K=freundlich_k,
                n=1 / one_over_n,
            )

        # Langmuir: columns E-F
        langmuir_k = sheet.cell(
            row=row,
            column=5,
        ).value

        q_m = sheet.cell(
            row=row,
            column=6,
        ).value

        if langmuir_k is not None and q_m is not None:
            chemical_isotherms["langmuir"] = LangmuirIsotherm(
                K=langmuir_k,
                q_m=q_m,
            )

        if chemical_isotherms:
            isotherms[normalized_name] = chemical_isotherms

    return isotherms


def load_input_file(
    parameter_file: str | Path,
    breakthrough_file: str | Path,
    breakthrough_sheet: str,
    isotherm_file: str | Path | None = None,
):
    """Load ReactorModels experimental input data."""
    parameter_workbook = openpyxl.load_workbook(
        parameter_file,
        data_only=True,
    )

    breakthrough_workbook = openpyxl.load_workbook(
        breakthrough_file,
        data_only=True,
    )

    parameter_sheet = parameter_workbook["parameters"]
    chemical_sheet = parameter_workbook["chemical"]

    selected_breakthrough_sheet = breakthrough_workbook[breakthrough_sheet]

    compound_names = [
        selected_breakthrough_sheet.cell(
            row=1,
            column=column_number,
        ).value
        for column_number in range(
            3,
            selected_breakthrough_sheet.max_column + 1,
        )
        if selected_breakthrough_sheet.cell(
            row=1,
            column=column_number,
        ).value
        is not None
    ]

    water, media, column = _load_properties(parameter_sheet)

    chemicals = _load_chemicals(
        chemical_sheet,
        compound_names,
    )

    breakthroughs = _load_breakthroughs(
        workbook=breakthrough_workbook,
        sheet_name=breakthrough_sheet,
        chemicals=chemicals,
        column=column,
    )

    isotherms = {}

    if isotherm_file is not None:
        isotherm_workbook = openpyxl.load_workbook(
            isotherm_file,
            data_only=True,
        )

        isotherm_sheet = isotherm_workbook.worksheets[0]

        isotherms = _load_isotherms(
            isotherm_sheet,
            compound_names,
        )

    return {
        "water": water,
        "media": media,
        "column": column,
        "chemicals": chemicals,
        "breakthroughs": breakthroughs,
        "isotherms": isotherms,
    }


def identify_curve_outliers(
    time, values, absolute_tolerance, relative_tolerance, window_size, max_outliers
):
    """Identify outliers using iterative local linear fits.

    1. Select a local window around each candidate point.
    2. Remove the candidate point individually and fit a linear trend
       to the remaining points.
    3. Calculate the candidate's deviation from the local trend.
    4. Calculate both an absolute and relative deviation.
    5. A point is considered an outlier if it exceeds either tolerance.
    6. Remove the point with the largest normalized deviation.
    7. Repeat using the reduced dataset.
    """
    time = np.asarray(time, dtype=float)
    values = np.asarray(values, dtype=float)

    if len(time) != len(values):
        raise ValueError("time and values must have the same length.")

    if window_size < 3:
        raise ValueError("window_size must be at least 3.")

    if absolute_tolerance < 0:
        raise ValueError("absolute_tolerance must be non-negative.")

    if relative_tolerance < 0:
        raise ValueError("relative_tolerance must be non-negative.")

    # Work with original indices so that the final outlier mask
    # corresponds to the original input arrays.
    remaining = list(range(len(time)))

    removed = []
    iteration_results = []

    for iteration in range(max_outliers):
        iteration_results = []

        half_window = window_size // 2

        # Evaluate every point as a potential outlier.
        for position, original_index in enumerate(remaining):

            # Determine the local window around the candidate.
            start = max(
                0,
                position - half_window,
            )
            end = start + window_size

            # Shift the window toward the beginning at end of the dataset.
            if end > len(remaining):
                end = len(remaining)
                start = end - window_size

            window_indices = remaining[start:end]

            # Remove the candidate from the points used to establish the local trend.
            fit_indices = [index for index in window_indices if index != original_index]

            t_fit = time[fit_indices]
            y_fit = values[fit_indices]

            # Fit the local linear trend.
            slope, intercept = np.polyfit(
                t_fit,
                y_fit,
                1,
            )

            # Predict the candidate using the local trend.
            predicted = np.maximum(
                slope * time[original_index] + intercept,
                0,
            )

            # Calculate the candidate's residual.
            residual = values[original_index] - predicted
            absolute_error = abs(residual)

            # Points below the detection threshold are not considered outliers.
            below_detection_threshold = values[original_index] < 0.01

            # Mixed absolute + relative tolerance.
            tolerance = absolute_tolerance + relative_tolerance * abs(predicted)

            is_outlier = not below_detection_threshold and absolute_error > tolerance

            iteration_results.append(
                {
                    "index": original_index,
                    "time": time[original_index],
                    "value": values[original_index],
                    "predicted": predicted,
                    "residual": residual,
                    "absolute_error": absolute_error,
                    "tolerance": tolerance,
                    "is_outlier": is_outlier,
                }
            )

        # Keep only points that exceed the mixed tolerance.
        candidates = [result for result in iteration_results if result["is_outlier"]]

        # Stop if no points exceed the tolerance.
        if not candidates:
            break

        # Determine how severely each candidate exceeds its tolerance.
        for result in candidates:
            result["violation_ratio"] = result["absolute_error"] / max(
                result["tolerance"],
                np.finfo(float).eps,
            )

        # Remove the point with the greatest violation.
        worst = max(
            candidates,
            key=lambda r: r["violation_ratio"],
        )

        removed.append(
            {
                "iteration": iteration + 1,
                "index": worst["index"],
                "time": worst["time"],
                "value": worst["value"],
                "predicted": worst["predicted"],
                "residual": worst["residual"],
                "absolute_error": worst["absolute_error"],
                "tolerance": worst["tolerance"],
                "violation_ratio": worst["violation_ratio"],
            }
        )

        # Remove the point using its original index.
        remaining.remove(worst["index"])

    # Construct the final outlier mask using the original indices.
    outlier = np.zeros(
        len(values),
        dtype=bool,
    )

    for result in removed:
        outlier[result["index"]] = True

    return outlier, iteration_results, removed
