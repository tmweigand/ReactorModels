"""Input_output.py."""

from pathlib import Path

import numpy as np
import openpyxl

from reactormodels.models import (
    FreundlichIsotherm,
    LangmuirIsotherm,
    LinearIsotherm,
)
from reactormodels.properties import (
    Breakthrough,
    Chemical,
    Column,
    Media,
    Water,
)


def normalize_name(name: str) -> str:
    """Normalize chemical names for matching."""
    return "".join(str(name).split()).casefold()


def _load_properties(sheet):
    """Create Water, Media, and Column objects."""
    water_parameters = {}
    media_parameters = {}
    column_parameters = {}

    sections = {
        "water": water_parameters,
        "media": media_parameters,
        "column": column_parameters,
    }

    current_section = None

    for parameter, value, *_ in sheet.iter_rows(values_only=True):
        if parameter is None:
            continue

        parameter_name = str(parameter).strip().rstrip(":")
        key = parameter_name.casefold()

        if key in sections:
            current_section = sections[key]
            continue

        if key in {
            "parameter",
            "parameters",
            "experiment_id",
        }:
            continue

        if current_section is not None and value is not None:
            current_section[parameter_name] = value

    water = Water(**water_parameters)
    media = Media(**media_parameters)

    column = Column(
        water=water,
        media=media,
        **column_parameters,
    )

    return water, media, column


def _load_chemicals(sheet, compound_names):
    """Create Chemical objects used in the breakthrough data."""
    chemicals = {}

    normalized_compounds = {normalize_name(name) for name in compound_names}

    for column_number in range(2, sheet.max_column + 1):
        chemical_name = sheet.cell(
            row=1,
            column=column_number,
        ).value

        if chemical_name is None:
            continue

        normalized_name = normalize_name(chemical_name)

        if normalized_name not in normalized_compounds:
            continue

        chemical_parameters = {}

        for row in range(2, sheet.max_row + 1):
            parameter_name = sheet.cell(
                row=row,
                column=1,
            ).value

            value = sheet.cell(
                row=row,
                column=column_number,
            ).value

            if parameter_name is None or value is None:
                continue

            chemical_parameters[str(parameter_name).strip()] = value

        chemicals[normalized_name] = Chemical(
            name=str(chemical_name).strip(),
            **chemical_parameters,
        )

    return chemicals


def _load_breakthroughs(
    workbook,
    sheet_name,
    chemicals,
    column,
):
    """Create Breakthrough objects."""
    parameter_sheet = workbook["breakthrough_parameters"]
    feed_sheet = workbook["feed_concentration"]
    selected_sheet = workbook[sheet_name]
    initial_sheet = workbook["initial_concentration"]

    breakthrough_parameters = {}

    for parameter, value, *_ in parameter_sheet.iter_rows(values_only=True):
        if parameter is None or value is None:
            continue

        parameter_name = str(parameter).strip()

        if parameter_name.casefold() in {
            "breakthrough",
            "parameter",
            "parameters",
        }:
            continue

        breakthrough_parameters[parameter_name] = value

    feed_columns = {
        normalize_name(feed_sheet.cell(1, column_number).value): column_number
        for column_number in range(3, feed_sheet.max_column + 1)
        if feed_sheet.cell(1, column_number).value is not None
    }

    initial_columns = {
        normalize_name(initial_sheet.cell(1, column_number).value): column_number
        for column_number in range(
            3,
            initial_sheet.max_column + 1,
        )
        if initial_sheet.cell(1, column_number).value is not None
    }

    breakthroughs = {}

    for column_number in range(
        3,
        selected_sheet.max_column + 1,
    ):
        chemical_name = selected_sheet.cell(
            row=1,
            column=column_number,
        ).value

        if chemical_name is None:
            continue

        normalized_name = normalize_name(chemical_name)

        if normalized_name not in chemicals:
            raise ValueError(f"{chemical_name!r} is not in the chemical sheet.")

        selected_concentrations = []
        bed_volumes = []
        time = []

        for row in range(2, selected_sheet.max_row + 1):
            concentration = selected_sheet.cell(
                row=row,
                column=column_number,
            ).value

            if concentration is None:
                continue

            selected_concentrations.append(concentration)

            bed_volumes.append(
                selected_sheet.cell(
                    row=row,
                    column=1,
                ).value
            )

            time.append(
                selected_sheet.cell(
                    row=row,
                    column=2,
                ).value
            )

        feed_column = feed_columns[normalized_name]

        feed_concentrations = [
            feed_sheet.cell(
                row=row,
                column=feed_column,
            ).value
            for row in range(2, feed_sheet.max_row + 1)
            if feed_sheet.cell(
                row=row,
                column=feed_column,
            ).value
            is not None
        ]

        breakthrough_kwargs = {
            "chemical": chemicals[normalized_name],
            "column": column,
            "feed_concentrations": np.asarray(
                feed_concentrations,
                dtype=float,
            ),
            "bed_volumes": np.asarray(
                bed_volumes,
                dtype=float,
            ),
            "time": np.asarray(
                time,
                dtype=float,
            ),
            **breakthrough_parameters,
        }

        if sheet_name == "effluent_concentration":
            breakthrough_kwargs["effluent_concentrations"] = np.asarray(
                selected_concentrations,
                dtype=float,
            )

        if normalized_name in initial_columns:
            initial_concentration = initial_sheet.cell(
                row=2,
                column=initial_columns[normalized_name],
            ).value

            if initial_concentration is not None:
                breakthrough_kwargs["initial_concentration"] = initial_concentration

        breakthroughs[normalized_name] = Breakthrough(**breakthrough_kwargs)

    return breakthroughs


def _load_isotherms(
    sheet,
    compound_names,
):
    """Create isotherm objects from supplied parameters."""
    isotherms = {}

    normalized_compounds = {normalize_name(name) for name in compound_names}

    for column_number in range(2, sheet.max_column + 1):
        chemical_name = sheet.cell(
            row=2,
            column=column_number,
        ).value

        if chemical_name is None:
            continue

        normalized_name = normalize_name(chemical_name)

        if normalized_name not in normalized_compounds:
            continue

        chemical_isotherms = {}

        # Linear
        linear_k = sheet.cell(
            row=4,
            column=column_number,
        ).value

        if linear_k is not None:
            chemical_isotherms["linear"] = LinearIsotherm(K=linear_k)

        # Freundlich
        freundlich_k = sheet.cell(
            row=6,
            column=column_number,
        ).value

        one_over_n = sheet.cell(
            row=7,
            column=column_number,
        ).value

        if freundlich_k is not None and one_over_n is not None:
            chemical_isotherms["freundlich"] = FreundlichIsotherm(
                K=freundlich_k,
                n=1 / one_over_n,
            )

        # Langmuir
        langmuir_k = sheet.cell(
            row=9,
            column=column_number,
        ).value

        q_m = sheet.cell(
            row=10,
            column=column_number,
        ).value

        if langmuir_k is not None and q_m is not None:
            chemical_isotherms["langmuir"] = LangmuirIsotherm(
                K=langmuir_k,
                q_m=q_m,
            )

        if chemical_isotherms:
            isotherms[normalized_name] = chemical_isotherms

    return isotherms


def load_input_file(
    parameter_file: str | Path,
    breakthrough_file: str | Path,
    breakthrough_sheet: str,
    isotherm_file: str | Path | None = None,
):
    """Load ReactorModels experimental input data."""
    parameter_workbook = openpyxl.load_workbook(
        parameter_file,
        data_only=True,
    )

    breakthrough_workbook = openpyxl.load_workbook(
        breakthrough_file,
        data_only=True,
    )

    parameter_sheet = parameter_workbook["parameters"]
    chemical_sheet = parameter_workbook["chemical"]

    selected_breakthrough_sheet = breakthrough_workbook[breakthrough_sheet]

    compound_names = [
        selected_breakthrough_sheet.cell(
            row=1,
            column=column_number,
        ).value
        for column_number in range(
            3,
            selected_breakthrough_sheet.max_column + 1,
        )
        if selected_breakthrough_sheet.cell(
            row=1,
            column=column_number,
        ).value
        is not None
    ]

    water, media, column = _load_properties(parameter_sheet)

    chemicals = _load_chemicals(
        chemical_sheet,
        compound_names,
    )

    breakthroughs = _load_breakthroughs(
        workbook=breakthrough_workbook,
        sheet_name=breakthrough_sheet,
        chemicals=chemicals,
        column=column,
    )

    isotherms = {}

    if isotherm_file is not None:
        isotherm_workbook = openpyxl.load_workbook(
            isotherm_file,
            data_only=True,
        )

        isotherm_sheet = isotherm_workbook.worksheets[0]

        isotherms = _load_isotherms(
            isotherm_sheet,
            compound_names,
        )

    return {
        "water": water,
        "media": media,
        "column": column,
        "chemicals": chemicals,
        "breakthroughs": breakthroughs,
        "isotherms": isotherms,
    }
